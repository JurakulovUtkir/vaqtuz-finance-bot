"""To'lov ma'lumotlarini Excel (.xlsx) ga chiqarish.

Eng muhim varaq — "Kanal narxlari": qatorlar kanallar, ustunlar oylar,
katakda o'rtacha post narxi. Shu jadval "avval qancha to'lagandik, hozir
qancha to'layapmiz" degan savolga bir qarashda javob beradi.
"""

from __future__ import annotations

import io
from typing import Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.db.models import PaymentRequest
from app.domain.analytics import (
    by_channel,
    by_month,
    by_project,
    by_week,
    channel_month_matrix,
    week_label,
)
from app.domain.formatting import format_resurs

MONEY = "#,##0"
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF")
TOTAL_FONT = Font(bold=True)


def _write_header(sheet: Worksheet, headers: Sequence[str], row: int = 1) -> None:
    for column, title in enumerate(headers, start=1):
        cell = sheet.cell(row=row, column=column, value=title)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.freeze_panes = sheet.cell(row=row + 1, column=1)


def _autosize(sheet: Worksheet, minimum: int = 10, maximum: int = 42) -> None:
    for column in sheet.columns:
        longest = max((len(str(cell.value)) for cell in column if cell.value is not None), default=0)
        letter = get_column_letter(column[0].column)
        sheet.column_dimensions[letter].width = min(max(longest + 2, minimum), maximum)


def _money_column(sheet: Worksheet, columns: Sequence[int], first_row: int = 2) -> None:
    for row in sheet.iter_rows(min_row=first_row):
        for cell in row:
            if cell.column in columns:
                cell.number_format = MONEY


def _sheet_summary(book: Workbook, requests: Sequence[PaymentRequest], title: str) -> None:
    sheet = book.active
    sheet.title = "Umumiy"

    paid = [r for r in requests if r.is_paid]
    pending = [r for r in requests if not r.is_paid]

    sheet["A1"] = title
    sheet["A1"].font = Font(bold=True, size=14)

    rows = [
        ("Jami so'rovlar", len(requests)),
        ("Jami summa (so'ralgan)", sum(r.summa_value for r in requests)),
        ("To'langan (soni)", len(paid)),
        ("To'langan (summa)", sum(r.summa_value for r in paid)),
        ("Kutilmoqda (soni)", len(pending)),
        ("Kutilmoqda (summa)", sum(r.summa_value for r in pending)),
        ("Haqiqiy o'tkazilgan", sum(r.effective_paid for r in paid)),
        ("Jami komissiya", sum(r.komissiya for r in paid)),
        ("Kanallar soni", len(by_channel(requests))),
        ("Mijozlar soni", len(by_project(requests))),
    ]
    for index, (label, value) in enumerate(rows, start=3):
        sheet.cell(row=index, column=1, value=label).font = TOTAL_FONT
        cell = sheet.cell(row=index, column=2, value=value)
        if "soni" not in label and "Kanallar" not in label and "Mijozlar" not in label:
            cell.number_format = MONEY
    _autosize(sheet)


def _sheet_channel_prices(book: Workbook, requests: Sequence[PaymentRequest]) -> None:
    matrix = channel_month_matrix(requests)
    sheet = book.create_sheet("Kanal narxlari")

    _write_header(sheet, ["Kanal", *matrix.months, "O'zgarish"])
    for row_index, channel in enumerate(matrix.channels, start=2):
        sheet.cell(row=row_index, column=1, value=channel)
        values: list[int] = []
        for column_index, month in enumerate(matrix.months, start=2):
            average = matrix.average(channel, month)
            if average is None:
                continue
            cell = sheet.cell(row=row_index, column=column_index, value=average)
            cell.number_format = MONEY
            values.append(average)

        change_cell = sheet.cell(row=row_index, column=len(matrix.months) + 2)
        if len(values) >= 2 and values[0]:
            ratio = (values[-1] - values[0]) / values[0]
            change_cell.value = ratio
            change_cell.number_format = "+0%;-0%;0%"
            change_cell.font = Font(color="C00000" if ratio > 0 else "008000", bold=True)
        else:
            change_cell.value = "—"
    _autosize(sheet, minimum=12)


def _sheet_group(
    book: Workbook,
    name: str,
    header: str,
    stats,
    *,
    relabel=None,
) -> None:
    sheet = book.create_sheet(name)
    _write_header(sheet, [header, "So'rovlar", "Jami summa", "O'rtacha"])
    for row_index, stat in enumerate(stats, start=2):
        label = relabel(stat.key) if relabel else stat.key
        sheet.cell(row=row_index, column=1, value=label)
        sheet.cell(row=row_index, column=2, value=stat.count)
        sheet.cell(row=row_index, column=3, value=stat.total)
        sheet.cell(row=row_index, column=4, value=stat.average)
    _money_column(sheet, (3, 4))
    _autosize(sheet)


def _sheet_raw(book: Workbook, requests: Sequence[PaymentRequest]) -> None:
    sheet = book.create_sheet("Barcha so'rovlar")
    _write_header(
        sheet,
        [
            "#", "Sana", "Kanal", "Mijoz", "So'ralgan", "Holat",
            "O'tkazilgan", "Komissiya", "Karta", "Kim so'radi", "To'langan sana",
        ],
    )
    for row_index, request in enumerate(requests, start=2):
        created = request.created
        sheet.cell(row=row_index, column=1, value=request.id)
        sheet.cell(row=row_index, column=2, value=created.strftime("%Y-%m-%d %H:%M") if created else "—")
        sheet.cell(row=row_index, column=3, value=format_resurs(request.resurs))
        sheet.cell(row=row_index, column=4, value=request.proyekt)
        sheet.cell(row=row_index, column=5, value=request.summa_value)
        sheet.cell(row=row_index, column=6, value="to'landi" if request.is_paid else "kutilmoqda")
        sheet.cell(row=row_index, column=7, value=request.effective_paid if request.is_paid else None)
        sheet.cell(row=row_index, column=8, value=request.komissiya or None)
        sheet.cell(row=row_index, column=9, value=request.karta)
        sheet.cell(row=row_index, column=10, value=request.requested_by)
        sheet.cell(row=row_index, column=11, value=(request.paid_at or "")[:16].replace("T", " "))
    _money_column(sheet, (5, 7, 8))
    _autosize(sheet)


def build_workbook(requests: Sequence[PaymentRequest], title: str) -> bytes:
    book = Workbook()
    _sheet_summary(book, requests, title)

    if requests:
        _sheet_channel_prices(book, requests)
        _sheet_group(book, "Kanallar", "Kanal", by_channel(requests))
        _sheet_group(book, "Oylar", "Oy", by_month(requests))
        _sheet_group(book, "Haftalar", "Hafta", by_week(requests), relabel=week_label)
        _sheet_group(book, "Mijozlar", "Mijoz", by_project(requests))
        _sheet_raw(book, requests)

    buffer = io.BytesIO()
    book.save(buffer)
    return buffer.getvalue()
