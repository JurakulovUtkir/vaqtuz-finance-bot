import io

from openpyxl import load_workbook

from app.db.models import STATUS_PAID
from app.export import build_workbook
from tests.factories import make_request


def _book(requests, title="Test"):
    return load_workbook(io.BytesIO(build_workbook(requests, title)))


def test_empty_export_has_only_summary():
    book = _book([])
    assert book.sheetnames == ["Umumiy"]
    assert book["Umumiy"]["A1"].value == "Test"


def test_all_sheets_present():
    book = _book([make_request(1)])
    assert book.sheetnames == [
        "Umumiy",
        "Kanal narxlari",
        "Kanallar",
        "Oylar",
        "Haftalar",
        "Mijozlar",
        "Barcha so'rovlar",
    ]


def test_summary_numbers():
    rows = [
        make_request(1, summa=60000, status=STATUS_PAID, actual=60600, komissiya=600),
        make_request(2, summa=40000),
    ]
    sheet = _book(rows)["Umumiy"]
    values = {sheet.cell(row=r, column=1).value: sheet.cell(row=r, column=2).value
              for r in range(3, 13)}
    assert values["Jami so'rovlar"] == 2
    assert values["Jami summa (so'ralgan)"] == 100000
    assert values["To'langan (soni)"] == 1
    assert values["Kutilmoqda (summa)"] == 40000
    assert values["Jami komissiya"] == 600


def test_channel_price_matrix_shows_months_side_by_side():
    """Asosiy varaq: kanal qatorda, oylar ustunda, o'zgarish foizda."""
    rows = [
        make_request(1, resurs="https://t.me/a", summa=1000000,
                     created_at="2026-06-10T10:00:00+05:00"),
        make_request(2, resurs="https://t.me/a", summa=1500000,
                     created_at="2026-07-10T10:00:00+05:00"),
    ]
    sheet = _book(rows)["Kanal narxlari"]

    headers = [c.value for c in sheet[1]]
    assert headers == ["Kanal", "2026-06", "2026-07", "O'zgarish"]

    assert sheet.cell(row=2, column=1).value == "@a"
    assert sheet.cell(row=2, column=2).value == 1000000
    assert sheet.cell(row=2, column=3).value == 1500000
    assert sheet.cell(row=2, column=4).value == 0.5  # +50%


def test_channel_with_single_month_has_no_change():
    rows = [make_request(1, resurs="https://t.me/a", created_at="2026-07-10T10:00:00+05:00")]
    sheet = _book(rows)["Kanal narxlari"]
    assert sheet.cell(row=2, column=3).value == "—"


def test_weeks_sheet_uses_readable_labels():
    rows = [make_request(1, created_at="2026-07-28T10:00:00+05:00")]  # seshanba
    sheet = _book(rows)["Haftalar"]
    assert sheet.cell(row=2, column=1).value == "27.07 - 02.08.2026"


def test_months_sheet_sorted_chronologically():
    rows = [
        make_request(1, summa=90000, created_at="2026-07-10T10:00:00+05:00"),
        make_request(2, summa=10000, created_at="2026-06-10T10:00:00+05:00"),
    ]
    sheet = _book(rows)["Oylar"]
    # Summa bo'yicha emas, vaqt bo'yicha tartiblangan bo'lishi kerak
    assert [sheet.cell(row=r, column=1).value for r in (2, 3)] == ["2026-06", "2026-07"]


def test_raw_sheet_has_every_request():
    rows = [make_request(i) for i in range(1, 6)]
    sheet = _book(rows)["Barcha so'rovlar"]
    assert sheet.max_row == 6  # sarlavha + 5
    assert [sheet.cell(row=r, column=1).value for r in range(2, 7)] == [1, 2, 3, 4, 5]


def test_raw_sheet_marks_status():
    rows = [make_request(1, status=STATUS_PAID, actual=60000), make_request(2)]
    sheet = _book(rows)["Barcha so'rovlar"]
    assert sheet.cell(row=2, column=6).value == "to'landi"
    assert sheet.cell(row=3, column=6).value == "kutilmoqda"
